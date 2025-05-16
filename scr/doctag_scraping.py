import os
from datetime import datetime
import regex as re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# --- Load Paths from Environment Variables ---
base_folder = os.environ.get("BASE_FOLDER", "data/test")
ens_syntax_file = os.environ.get("ENS_SYNTAX_FILE", "../data/ENS_Syntax.txt")
tags_path = os.environ.get("TAGS_PATH", "../data/Tags.xlsx")
docs_path = os.environ.get("DOCS_PATH", "../data/Docs.xlsx")
doc_tag_path = os.environ.get("DOC_TAG_PATH", "../data/Doc-Tag.xlsx")
error_log_file = os.path.join(base_folder, "error_log.txt")


# --- Logging ---
def log_error(message):
    with open(error_log_file, 'a', encoding='utf-8') as log_file:
        log_file.write(message + "\n")


# --- ENS Pattern Loader ---
def load_syntax_patterns(file_path):
    try:
        patterns = []
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if line.strip():
                    pattern = re.compile(
                        r"\b\d{2}-[A-Z]{2,4}-\d{3,5}[A-Z]?([-/][A-Z0-9]{1,3})*\b",
                        re.IGNORECASE | re.UNICODE | re.BESTMATCH
                    )
                    patterns.append(pattern)
        print(f"Loaded {len(patterns)} patterns from {file_path}")
        return patterns
    except Exception as e:
        log_error(f"Error loading ENS patterns: {e}")
        return []


# --- Pattern Matching ---
def matches_any_ens_mask(tag, ens_patterns):
    tag = tag.strip().replace("\u200b", "").replace("\ufeff", "")
    return any(pat.fullmatch(tag) for pat in ens_patterns)


# --- Slash Expansion ---
def expand_slash(tag_text):
    if '/' not in tag_text:
        return [tag_text]
    parts = tag_text.split('/')
    base = parts[0]
    match = re.match(r'^(.*?)-(\d+)([A-Za-z]*)$', base)
    if not match:
        return [tag_text]
    prefix, num_part, suffix = match.groups()
    expanded = [base]
    for part in parts[1:]:
        part = part.strip()
        if part.isdigit():
            expanded.append(f"{prefix}-{part}{suffix}")
        elif re.match(r'^[A-Za-z]+$', part):
            expanded.append(f"{prefix}-{num_part}{part}")
        else:
            expanded.append(f"{prefix}-{part}")
    return expanded


# --- XML Text Extraction ---
def extract_text_with_pages(xml_path):
    try:
        with open(xml_path, 'r', encoding='utf-16') as file:
            xml_text = file.read()
        page_pattern = re.compile(r'<page[^>]*>', re.IGNORECASE)
        pages = list(page_pattern.finditer(xml_text))
        extracted_data = []
        for i, page in enumerate(pages):
            start = page.end()
            end = pages[i + 1].start() if i + 1 < len(pages) else len(xml_text)
            content = re.sub(r'<[^>]+>', ' ', xml_text[start:end])
            content = re.sub(r'\s+', ' ', content).strip()
            extracted_data.append((i + 1, content))
        return extracted_data
    except Exception as e:
        log_error(f"Error reading {xml_path}: {e}")
        return []


# --- Tag Extraction w/ Pages ---
def extract_tags_with_pages(pages_data, ens_patterns):
    tags = []
    found = set()
    for page_num, text in pages_data:
        candidates = set(re.findall(r"\b[A-Z0-9]+(?:-[A-Z0-9]+)*(?:/[A-Z0-9]+)*\b", text))
        matched = {tag for tag in candidates if matches_any_ens_mask(tag, ens_patterns)}
        expanded = set()
        for tag in matched:
            expanded.update(expand_slash(tag)) if '/' in tag else expanded.add(tag)
        for tag in expanded:
            if (tag, page_num) not in found:
                tags.append((tag, page_num))
                found.add((tag, page_num))
    return tags


def load_docs_values(file_path):
    try:
        wb = load_workbook(file_path)
        return [str(cell).strip() for row in wb.active.iter_rows(values_only=True) for cell in row if cell]
    except Exception as e:
        log_error(f"Error loading Docs.xlsx: {e}")
        return []


def process_xml_file(xml_path, ens_patterns, docs_values):
    pages_data = extract_text_with_pages(xml_path)
    if not pages_data:
        return []
    extracted = extract_tags_with_pages(pages_data, ens_patterns)
    results = []
    for tag, page in extracted:
        if not any(tag in doc for doc in docs_values):
            results.append((tag, os.path.basename(xml_path), page))
    return results


def extract_document_no(filename):
    return filename.replace(".xml", "")


def load_doc_tag_values(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            tag_no, doc_no, action = row[:3]
            if doc_no and tag_no:
                base = doc_no.split("_")[0]
                mapping.setdefault(base, {})[tag_no] = action
        return mapping
    except Exception as e:
        log_error(f"Error loading Doc-Tag.xlsx: {e}")
        return {}


def load_tag_status(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        status_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            tag_no, _, status = row[:3]
            if tag_no:
                status_map[tag_no] = status or "Unidentified"
        return status_map
    except Exception as e:
        log_error(f"Error loading Tags.xlsx: {e}")
        return {}


def process_folder(folder_path, patterns, docs_values, doc_tag_map, tag_status_map):
    folder_name = os.path.basename(folder_path)
    output_xlsx = os.path.join(folder_path, f"{folder_name}-Doc-Tag-Scraping.xlsx")
    xml_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xml")]
    all_tags = []
    seen = set()

    for xml_file in xml_files:
        document_no = extract_document_no(os.path.basename(xml_file))
        base_doc_no = document_no.split("_")[0]
        tags = process_xml_file(xml_file, patterns, docs_values)
        tag_dict = {}
        for tag, _, page in tags:
            tag_dict.setdefault(tag, page)
        for tag, page in tag_dict.items():
            action = doc_tag_map.get(base_doc_no, {}).get(tag, "")
            status = tag_status_map.get(tag, "Unidentified")
            all_tags.append([tag, document_no, page, action, status])
            seen.add((base_doc_no, tag))

    # Append missing tags
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    missing = []
    for base_doc_no, tags_dict in doc_tag_map.items():
        if any(base_doc_no in doc for _, doc, *_ in all_tags):
            for tag_no, action in tags_dict.items():
                if (base_doc_no, tag_no) not in seen:
                    missing.append([tag_no, base_doc_no, "", action, ""])
    all_tags.extend(missing)

    if all_tags:
        wb = Workbook()
        ws = wb.create_sheet("Doc-Tag")
        wb.remove(wb.active)  # Remove default sheet

        headers = ["Tag No", "DocumentNo", "Page", "Action", "Status"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        hyperlink_template = '=HYPERLINK("{pdf}", "{label}")'
        for i, (tag, doc_no, page, action, status) in enumerate(all_tags, start=2):
            if page == "":
                ws.append([tag, doc_no, page, action, status])
            else:
                pdf_path = f"{folder_path}/{doc_no}.pdf"
                ws.append([tag, hyperlink_template.format(pdf=pdf_path, label=doc_no), page, action, status])
            if [tag, doc_no, page, action, status] in missing:
                for cell in ws[i]:
                    cell.fill = green_fill

        wb.save(output_xlsx)
        print(f"✅ Saved: {output_xlsx}")
    else:
        print(f"❌ No tags found in: {folder_path}")


def main():
    patterns = load_syntax_patterns(ens_syntax_file)
    if not patterns:
        print("⚠️ ENS patterns missing.")
        return

    docs = load_docs_values(docs_path)
    doc_tag_map = load_doc_tag_values(doc_tag_path)
    tag_status = load_tag_status(tags_path)

    for root, _, files in os.walk(base_folder):
        if any(f.endswith(".xml") for f in files):
            print(f"📂 Processing: {root}")
            process_folder(root, patterns, docs, doc_tag_map, tag_status)


if __name__ == "__main__":
    main()
